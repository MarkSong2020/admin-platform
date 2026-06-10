"""通用 Excel 机制单测（P5 Excel 阶段1）——往返一致 / 坏行不阻断 / 缺列头 / 类型漂移 / 行数上限。"""

from __future__ import annotations

import zipfile
from io import BytesIO

from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field

from admin_platform.excel import ExcelColumn, ExcelExporter, ExcelImporter

_COLUMNS = [
    ExcelColumn("name", "名称"),
    ExcelColumn("code", "编码"),
    ExcelColumn("count", "数量", required=False),
]


class _SampleRow(BaseModel):
    name: str = Field(min_length=1, max_length=10)
    code: str = Field(min_length=1)
    count: int = 0


def _manual_xlsx(rows: list[list[object]]) -> bytes:
    """手写 xlsx（含数字/空 cell），模拟用户真实 Excel（绕过 exporter 的 canonical 化）。"""
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    for row in rows:
        worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_roundtrip_export_then_import() -> None:
    content = ExcelExporter(_COLUMNS).write(
        [{"name": "甲", "code": "x1", "count": 5}, {"name": "乙", "code": "x2", "count": 0}]
    )
    result = ExcelImporter(_SampleRow, _COLUMNS).parse(content)
    assert result.errors == []
    assert [(r.data.name, r.data.code, r.data.count) for r in result.rows] == [
        ("甲", "x1", 5),
        ("乙", "x2", 0),
    ]
    assert [r.row for r in result.rows] == [2, 3]  # Excel 行号（含表头）保留


def test_bad_row_collected_not_blocking() -> None:
    # 第 2 行 name 空（必填失败），第 3 行正常 → 好行通过 + 坏行定位
    content = ExcelExporter(_COLUMNS).write(
        [{"name": "", "code": "x1"}, {"name": "ok", "code": "x2"}]
    )
    result = ExcelImporter(_SampleRow, _COLUMNS).parse(content)
    assert [r.data.name for r in result.rows] == ["ok"]
    assert len(result.errors) == 1
    assert result.errors[0].row == 2
    assert result.errors[0].column == "名称"
    assert result.errors[0].code == "VALIDATION"


def test_optional_empty_uses_default() -> None:
    # 非必填 count 空 cell → 省略 → Pydantic default 0（不误报 int 解析失败）
    content = _manual_xlsx([["名称", "编码", "数量"], ["甲", "x1", None]])
    result = ExcelImporter(_SampleRow, _COLUMNS).parse(content)
    assert result.errors == []
    assert result.rows[0].data.count == 0


def test_numeric_cell_canonical_to_str() -> None:
    # 编码列填数字 123（用户 Excel 数字格式）→ canonical 成 str "123"，不漂移
    content = _manual_xlsx([["名称", "编码", "数量"], ["item", 123, 5]])
    result = ExcelImporter(_SampleRow, _COLUMNS).parse(content)
    assert result.errors == []
    assert result.rows[0].data.code == "123"


def test_missing_column_header() -> None:
    # 导出只含「名称」列，导入期望「编码」→ MISSING_COLUMN
    content = ExcelExporter([ExcelColumn("name", "名称")]).write([{"name": "甲"}])
    result = ExcelImporter(_SampleRow, _COLUMNS).parse(content)
    assert result.rows == []
    assert any(e.code == "MISSING_COLUMN" and e.column == "编码" for e in result.errors)


def test_empty_data_no_error() -> None:
    # 只表头无数据行 → 空结果，无错误（非 EMPTY，EMPTY 是连表头都没）
    content = ExcelExporter(_COLUMNS).write([])
    result = ExcelImporter(_SampleRow, _COLUMNS).parse(content)
    assert result.rows == []
    assert result.errors == []


def test_completely_empty_workbook() -> None:
    content = _manual_xlsx([])  # 无任何行
    result = ExcelImporter(_SampleRow, _COLUMNS).parse(content)
    assert any(e.code == "EMPTY" for e in result.errors)


def test_max_rows_exceeded() -> None:
    content = ExcelExporter(_COLUMNS).write([{"name": f"n{i}", "code": f"c{i}"} for i in range(5)])
    result = ExcelImporter(_SampleRow, _COLUMNS, max_rows=3).parse(content)
    assert len(result.rows) == 3
    assert any(e.code == "TOO_MANY_ROWS" for e in result.errors)


def test_blank_rows_skipped() -> None:
    content = _manual_xlsx(
        [["名称", "编码", "数量"], ["甲", "x1", 1], [None, None, None], ["乙", "x2", 2]]
    )
    result = ExcelImporter(_SampleRow, _COLUMNS).parse(content)
    assert [r.data.name for r in result.rows] == ["甲", "乙"]  # 全空行跳过
    assert result.errors == []


# ---- 对抗审查回归：formula injection / 非法 xlsx ---------------------------


def test_writer_escapes_formula_injection() -> None:
    # 导出 cell 以 =/+/@ 开头 → 前缀单引号文本化，data_type='s'（字符串非公式 'f'），防 Excel 执行
    content = ExcelExporter(_COLUMNS).write([{"name": "=1+1", "code": "+cmd", "count": "@x"}])
    sheet = load_workbook(BytesIO(content)).active
    assert sheet is not None
    assert sheet.cell(row=2, column=1).data_type == "s"
    assert sheet.cell(row=2, column=1).value == "'=1+1"
    assert sheet.cell(row=2, column=2).value == "'+cmd"
    assert sheet.cell(row=2, column=3).value == "'@x"


def test_writer_normal_value_not_escaped() -> None:
    content = ExcelExporter(_COLUMNS).write([{"name": "工程师", "code": "eng", "count": "5"}])
    sheet = load_workbook(BytesIO(content)).active
    assert sheet is not None
    assert sheet.cell(row=2, column=1).value == "工程师"  # 正常值无前导单引号，往返一致


def test_invalid_xlsx_returns_business_error() -> None:
    # 损坏/非 xlsx 内容 → INVALID_FILE 业务错误，不冒泡 500
    result = ExcelImporter(_SampleRow, _COLUMNS).parse(b"this is not a zip/xlsx at all")
    assert result.rows == []
    assert any(e.code == "INVALID_FILE" for e in result.errors)


def test_writer_escapes_newline_prefix_formula() -> None:
    # \n 开头的公式也转义（R2 闭合第二轮发现的理论缺口）
    content = ExcelExporter(_COLUMNS).write([{"name": "\n=1+1", "code": "x", "count": "1"}])
    sheet = load_workbook(BytesIO(content)).active
    assert sheet is not None
    assert sheet.cell(row=2, column=1).data_type == "s"
    assert sheet.cell(row=2, column=1).value == "'\n=1+1"


def test_writer_strips_illegal_control_chars() -> None:
    # 含 openpyxl 非法控制字符（\x0b 垂直制表）的值 → writer 剥除，导出不抛 IllegalCharacterError。
    # 对抗审查 R5 存储型 DoS 兜底：低权限用户投毒一行即可让全表导出对所有人永久 500，writer 层剥除
    # 让导出在结构上不可能因非法字符失败。
    content = ExcelExporter(_COLUMNS).write([{"name": "\x0b评估专员", "code": "c1", "count": "1"}])
    sheet = load_workbook(BytesIO(content)).active
    assert sheet is not None
    assert sheet.cell(row=2, column=1).value == "评估专员"  # 非法字符已剥除


def test_writer_strips_illegal_char_then_escapes_exposed_formula() -> None:
    # 剥除前导非法字符后若暴露出公式触发字符，仍正确转义（剥除发生在 formula 判断之前）。
    content = ExcelExporter(_COLUMNS).write([{"name": "\x00=1+1", "code": "c", "count": "1"}])
    sheet = load_workbook(BytesIO(content)).active
    assert sheet is not None
    assert sheet.cell(row=2, column=1).data_type == "s"
    assert sheet.cell(row=2, column=1).value == "'=1+1"


def test_valid_zip_non_xlsx_returns_business_error() -> None:
    # 有效 zip 但非 xlsx（缺 workbook part）→ openpyxl 抛 KeyError（非 BadZipFile）→ reader 的宽
    # ``except Exception`` 仍捕获成 INVALID_FILE，不退化 500。守护该回退：收窄到具体异常类型会漏
    # openpyxl 对此类文件的多类拒绝异常（KeyError/OSError），此测试锁住任何收窄改动会直接红。
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w") as archive:
        archive.writestr("hello.txt", "not an xlsx workbook")
    result = ExcelImporter(_SampleRow, _COLUMNS).parse(buffer.getvalue())
    assert result.rows == []
    assert any(e.code == "INVALID_FILE" for e in result.errors)


def test_writer_strips_noncharacter() -> None:
    # U+FFFE/U+FFFF 非字符不被 openpyxl ILLEGAL 集捕获（worksheet.append 不抛），却会生成损坏 .xlsx
    # （XML 1.0 Char 上限 U+FFFD）——writer 剥除后导出文件可正常重开（对抗审查 R5 skeptic 扩面闭合）。
    content = ExcelExporter(_COLUMNS).write(
        [{"name": chr(0xFFFE) + "甲", "code": chr(0xFFFF) + "x1", "count": "1"}]
    )
    sheet = load_workbook(BytesIO(content)).active  # 能重开即未损坏
    assert sheet is not None
    assert sheet.cell(row=2, column=1).value == "甲"
    assert sheet.cell(row=2, column=2).value == "x1"


def test_reader_invalid_xml_char_in_sheet_returns_business_error() -> None:
    # sheet XML 内含 XML 非法字符（U+FFFE 的 UTF-8 字节）→ load_workbook 成功但 iter_rows 解析抛 →
    # reader 迭代期 except 兜底转 INVALID_FILE，不退化 500（对抗审查 R5 skeptic：此前 reader 的 except
    # 只包 load_workbook，iter_rows 期间异常漏成 500，与 docstring「损坏 xlsx 转 INVALID_FILE」不符）。
    good = ExcelExporter(_COLUMNS).write([{"name": "甲", "code": "x1", "count": "1"}])
    buf_out = BytesIO()
    with zipfile.ZipFile(BytesIO(good)) as zin, zipfile.ZipFile(buf_out, "w") as zout:
        for item in zin.namelist():
            data = zin.read(item)
            if item.startswith("xl/worksheets/"):
                data = data.replace(b"x1", b"x1\xef\xbf\xbe")  # 注入 U+FFFE（UTF-8）破坏 sheet XML
            zout.writestr(item, data)
    result = ExcelImporter(_SampleRow, _COLUMNS).parse(buf_out.getvalue())
    assert result.rows == []
    assert any(e.code == "INVALID_FILE" for e in result.errors)
