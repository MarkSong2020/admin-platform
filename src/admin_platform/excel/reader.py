"""Excel 读取 + 逐行校验（openpyxl read-only 流式，无 domain 知识）。

canonical 文本化：每个 cell 一律 ``str`` 化并去首尾空白，空 cell→``""``（让 Pydantic 必填校验
捕获，非静默 null）。⚠️ **限制**（对抗审查 P2）：openpyxl 在此之前已按 cell 格式读类型——数字
格式的 ``007`` 被读成 int 7（前导零丢）、``1E5`` 读成 float。要保前导零/大数字，**编码列须在
Excel 存为文本格式**（导入模板应预设）。canonical 只保证「读到什么 → str 化一致」，不能从
numeric cell 逆转用户原本看到的文本。坏行不阻断后续行，全量收集 RowError；非法/损坏 xlsx 转
INVALID_FILE 业务错误不冒泡 500。

zip bomb / inflated XML DoS 防护（``parse`` 在交给 openpyxl 解压前先查中央目录）：xlsx 本质是
zip，``excel_max_upload_size_bytes`` 只限压缩后体积——小压缩包可膨胀出巨大 sharedStrings.xml /
sheet XML，openpyxl 解压解析时 CPU/内存 DoS（行数上限在解析**之后**才生效，挡不住）。预检三重：
总声明解压大小 / 条目数 / 单条目压缩比，任一超限抛 ``ExcelTooLargeError``（叶子异常，domain 映 413）。
⚠️ **诚实标注残留**：中央目录的 ``ZipInfo.file_size`` 是写入方声明值、攻击者可控——精心构造的 bomb
可**少报** file_size 而实际解压更大，本预检挡不住「谎报 file_size」。三重启发式拦住绝大多数
naive/常见 bomb（朴素 inflate、海量条目、极端压缩比），但不是对抗谎报 file_size 的 bulletproof
防御——彻底免疫需 bounded streaming 解压（边解压边记账，超阈值即中止），属更深排期。
"""

from __future__ import annotations

import zipfile
from collections.abc import Iterator, Sequence
from io import BytesIO

from openpyxl import load_workbook
from pydantic import BaseModel, ValidationError

from admin_platform.excel.schemas import (
    ExcelColumn,
    ExcelTooLargeError,
    ImportResult,
    ParsedRow,
    RowError,
)

# 默认阈值（permissive 但安全）：调用方（domain）通常从 settings 显式传入；这些默认是给未传值的
# 其他叶子使用者兜底，且要宽到不误伤正常 xlsx（正常 xlsx 解压几 MB、条目几十个、压缩比 < ~20x）。
_DEFAULT_MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024  # 100MB
_DEFAULT_MAX_ENTRIES = 512
_DEFAULT_MAX_RATIO = 100


class ExcelImporter[T: BaseModel]:
    def __init__(  # noqa: PLR0913 —— 三个 zip bomb 阈值均为命名 kwargs（带默认），调用方可只传需要的
        self,
        schema: type[T],
        columns: Sequence[ExcelColumn],
        *,
        max_rows: int = 10000,
        max_uncompressed_bytes: int = _DEFAULT_MAX_UNCOMPRESSED_BYTES,
        max_zip_entries: int = _DEFAULT_MAX_ENTRIES,
        max_compression_ratio: int = _DEFAULT_MAX_RATIO,
    ) -> None:
        self._schema = schema
        self._columns = list(columns)
        self._max_rows = max_rows
        self._max_uncompressed_bytes = max_uncompressed_bytes
        self._max_zip_entries = max_zip_entries
        self._max_compression_ratio = max_compression_ratio

    def _guard_zip_bomb(self, content: bytes) -> None:
        """解压前查中央目录：三重阈值任一超限抛 ``ExcelTooLargeError``。

        ``BadZipFile`` / 非 zip 不在此处理——交回 ``parse`` 的 ``load_workbook`` except 走
        INVALID_FILE 业务错误（不退化 500）。仅读中央目录元数据，**不解压**任何条目。
        """
        try:
            with zipfile.ZipFile(BytesIO(content)) as archive:
                infos = archive.infolist()
                if len(infos) > self._max_zip_entries:
                    raise ExcelTooLargeError(
                        f"zip 条目数 {len(infos)} 超上限 {self._max_zip_entries}"
                    )
                total = 0
                for info in infos:
                    total += info.file_size
                    if total > self._max_uncompressed_bytes:
                        raise ExcelTooLargeError(
                            f"声明解压总大小超上限 {self._max_uncompressed_bytes} 字节"
                        )
                    ratio = info.file_size / max(info.compress_size, 1)
                    if ratio > self._max_compression_ratio:
                        raise ExcelTooLargeError(
                            f"条目 {info.filename!r} 压缩比 {ratio:.0f}x 超上限 "
                            f"{self._max_compression_ratio}x"
                        )
        except zipfile.BadZipFile:
            return  # 非 zip / 坏 zip → 不在此拦，交 load_workbook 走 INVALID_FILE

    def parse(self, content: bytes) -> ImportResult[T]:
        # 解压前预检（zip bomb / inflated XML），超限抛 ExcelTooLargeError
        self._guard_zip_bomb(content)
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:  # 解析不可信 xlsx：任何解析失败转业务错误不退化 500（收窄会漏 openpyxl 对非法文件的多类拒绝异常 KeyError/OSError/ParseError，R3 验证证实）
            return ImportResult(
                rows=[],
                errors=[RowError(1, None, "INVALID_FILE", f"无法解析 Excel: {type(exc).__name__}")],
            )
        try:
            worksheet = workbook.active
            if worksheet is None:
                return ImportResult(rows=[], errors=[RowError(1, None, "EMPTY", "空工作簿")])
            rows_iter = worksheet.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if header is None:
                return ImportResult(rows=[], errors=[RowError(1, None, "EMPTY", "空文件")])
            header_map = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
            missing = [c.header for c in self._columns if c.header not in header_map]
            if missing:
                return ImportResult(
                    rows=[],
                    errors=[RowError(1, h, "MISSING_COLUMN", f"缺列头: {h}") for h in missing],
                )

            return self._parse_rows(rows_iter, header_map)
        except Exception as exc:
            # 行迭代期 openpyxl 解析异常（损坏 XML 的 ParseError / OSError 等）→ INVALID_FILE 不退化
            # 500：逐行 Pydantic 错误已在内层 except 收集，不会漏到此处。对抗审查 R5 skeptic：iter_rows
            # 期间 ParseError（如含 U+FFFE 的文件）此前漏成 500，与本文 docstring 声称不符。
            return ImportResult(
                rows=[],
                errors=[RowError(1, None, "INVALID_FILE", f"无法解析 Excel: {type(exc).__name__}")],
            )
        finally:
            workbook.close()

    def _parse_rows(
        self, rows_iter: Iterator[tuple[object, ...] | None], header_map: dict[str, int]
    ) -> ImportResult[T]:
        rows: list[ParsedRow[T]] = []
        errors: list[RowError] = []
        for row_idx, raw in enumerate(rows_iter, start=2):
            if row_idx - 1 > self._max_rows:
                errors.append(
                    RowError(row_idx, None, "TOO_MANY_ROWS", f"超过最大行数 {self._max_rows}")
                )
                break
            if raw is None or all(cell is None for cell in raw):
                continue  # 跳过全空行
            data: dict[str, str] = {}
            for col in self._columns:
                idx = header_map[col.header]
                value = _canonical(raw[idx] if idx < len(raw) else None)
                # 非必填字段空值省略 → 让 Pydantic 用 default（避免空串喂给 int/枚举字段误报）。
                if value == "" and not col.required:
                    continue
                data[col.field] = value
            try:
                instance = self._schema.model_validate(data)
            except ValidationError as exc:
                errors.extend(self._row_errors(row_idx, exc))
            else:
                rows.append(ParsedRow(row_idx, instance))
        return ImportResult(rows=rows, errors=errors)

    def _row_errors(self, row_idx: int, exc: ValidationError) -> list[RowError]:
        result: list[RowError] = []
        for err in exc.errors():
            field_name = str(err["loc"][0]) if err["loc"] else None
            header = next((c.header for c in self._columns if c.field == field_name), None)
            result.append(RowError(row_idx, header, "VALIDATION", err["msg"]))
        return result


def _canonical(value: object) -> str:
    """cell → canonical 文本（去首尾空白；None→空串）。"""
    return "" if value is None else str(value).strip()
